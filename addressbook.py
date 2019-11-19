import openpyxl
from openpyxl import load_workbook

import os

class Contact:
    def __init__(self,name,phone,email,addr):
        self.name = name
        self.phone = phone
        self.email = email
        self.addr = addr

    def print_info(self):
        print("\n")
        print("Name: ", self.name)
        print("Phone Number: ", self.phone)
        print("E-mail: ", self.email)
        print("Address: ", self.addr)
        print("\n")

def print_contact(contact_list):
    for contact in contact_list:
        contact.print_info()

def print_contact_renew():
    load_wb = load_workbook('addressbook.xlsx')
    load_ws = load_wb['Sheet']

    all_value = []
    for row in sheet1.iter_rows():
        a = []
        for cell in row:
            a.append(cell.value)
        all_value.append(a)


# def set_contact():
#     name = input("Name  : ")
#     phone = input("PH   : ")
#     email = input("Email: ")
#     addr = input("Addr  : ")
#     contact = Contact(name,phone,email,addr)
#     return contact

def set_contact_renew():
    #addressbook.xlsx 파일읽기
    # wb = openpyxl.load_workbook('addressbook.xlsx')
    # sheet1 = wb.active
    name = input("Name      : ")
    phone = input("PH       : ")
    email = input("Email    : ")
    addr = input("Address   : ")
    sheet1.append([name, phone, email, addr])
    print("\n--입력이 완료되었습니다--\n")
    wb.save('addressbook.xlsx')
    contact = Contact(name, phone, email, addr)
    return contact

def delete_contact(contact_list,name):
    for i, contact in enumerate(contact_list):
        if contact.name == name:
            del contact_list[i]

def print_menu():
    print("1. 연락처 입력")
    print("2. 연락처 출력")
    print("3. 연락처 삭제")
    print("4. 종료")
    menu = input("메뉴선택: ")
    return int(menu)

def run():
    contact_list = []
    while 1:
        menu = print_menu()
        if menu == 1:
            contact = set_contact_renew()
            contact_list.append(contact)
        elif menu == 2:
            print_contact_renew()
        elif menu == 3:
            name = input("Delete Name : ")
            delete_contact(contact_list, name)
        elif menu == 4:
            break
        else:
            print("Error")

#초기 엑셀 파일 생성


#C:\Users\Mark\Desktop\GitHub\mySTG
path = "/Users/Mark/Desktop/GitHub/mySTG"
file_names = os.listdir(path)

for file_name in file_names:
    if file_name == 'addressbook.xlsx':
        wb = openpyxl.load_workbook('addressbook.xlsx')
        sheet1 = wb.active
    else:
        wb = openpyxl.Workbook()
        wb.save('addressbook.xlsx')
        sheet1 = wb.active

if __name__ == "__main__":
    run()